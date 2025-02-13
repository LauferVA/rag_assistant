# rag_assistant/file_parsers.py

import os
import logging

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------
# Optional Imports: wrap them in try/except so we don't force users
# to install everything if they only need minimal text indexing.
# ---------------------------------------------------------------------
try:
    import docx  # For .docx
except ImportError:
    docx = None

try:
    import fitz  # PyMuPDF for PDFs
except ImportError:
    fitz = None

try:
    import PyPDF2
except ImportError:
    PyPDF2 = None

try:
    import openpyxl  # For .xlsx
except ImportError:
    openpyxl = None

try:
    import odf.text, odf.teletype, odf.opendocument  # For .odt
except ImportError:
    odf = None

import json


def extract_text_from_file(filepath):
    """
    Attempt to extract text from the given filepath, based on its extension.
    Return an empty string if parsing fails or is not supported.
    """
    ext = os.path.splitext(filepath)[1].lower()

    # Simple text-based extensions
    if ext in (
        ".txt", ".md", ".py", ".json", ".log", ".csv", ".tsv", ".xml",
        ".yaml", ".yml", ".html", ".htm", ".css", ".js", ".jsx", ".ts",
        ".tsx", ".sh", ".cmd", ".ps1", ".swift", ".kt", ".go", ".rs",
        ".lua", ".pl", ".r", ".m", ".vb", ".cs", ".asm", ".dart", ".php",
        ".rb", ".sql"
    ):
        return _read_plaintext(filepath)

    # DOCX
    if ext == ".docx":
        return _read_docx(filepath) if docx else ""

    # Older .doc or .rtf files: not fully supported in Python by default.
    if ext in (".doc", ".rtf"):
        logger.warning("Native reading not implemented for %s. Consider external tools.", ext)
        return ""

    # PDF
    if ext == ".pdf":
        if fitz:
            return _read_pdf_pymupdf(filepath)
        elif PyPDF2:
            return _read_pdf_pypdf2(filepath)
        else:
            logger.warning("No PDF library installed; cannot parse %s", filepath)
            return ""

    # ODT
    if ext == ".odt":
        return _read_odt(filepath) if odf else ""

    # XLSX, XLS, ODS
    if ext in (".xlsx", ".xlsm", ".xls", ".ods"):
        return _read_excel(filepath) if openpyxl else ""

    # PPT, PPTX, ODP
    # (python-pptx can handle .pptx, but not .ppt or .odp well)
    if ext in (".pptx", ".ppt", ".odp"):
        logger.warning("PPT/ODP parsing not implemented. Skipping %s", filepath)
        return ""

    # IPYNB
    if ext == ".ipynb":
        return _read_ipynb(filepath)

    # Fallback
    logger.debug("No parsing rule for %s; skipping.", filepath)
    return ""


def _read_plaintext(filepath):
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            return f.read()
    except Exception as e:
        logger.error("Error reading plaintext file %s: %s", filepath, e)
        return ""


def _read_docx(filepath):
    if not docx:
        return ""
    try:
        doc = docx.Document(filepath)
        paragraphs = [p.text for p in doc.paragraphs]
        return "\n".join(paragraphs)
    except Exception as e:
        logger.error("Error reading DOCX %s: %s", filepath, e)
        return ""


def _read_pdf_pymupdf(filepath):
    try:
        text_pages = []
        with fitz.open(filepath) as pdf_doc:
            for page in pdf_doc:
                text_pages.append(page.get_text())
        return "\n".join(text_pages)
    except Exception as e:
        logger.error("Error reading PDF (PyMuPDF) %s: %s", filepath, e)
        return ""


def _read_pdf_pypdf2(filepath):
    try:
        text_pages = []
        with open(filepath, "rb") as f:
            pdf = PyPDF2.PdfReader(f)
            for page_num in range(len(pdf.pages)):
                text_pages.append(pdf.pages[page_num].extract_text() or "")
        return "\n".join(text_pages)
    except Exception as e:
        logger.error("Error reading PDF (PyPDF2) %s: %s", filepath, e)
        return ""


def _read_odt(filepath):
    try:
        doc = odf.opendocument.load(filepath)
        text_elements = doc.getElementsByType(odf.text.P)
        paragraphs = [odf.teletype.extractText(elem) for elem in text_elements]
        return "\n".join(paragraphs)
    except Exception as e:
        logger.error("Error reading ODT %s: %s", filepath, e)
        return ""


def _read_excel(filepath):
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        text_chunks = []
        for name in wb.sheetnames:
            sheet = wb[name]
            for row in sheet.iter_rows(values_only=True):
                row_str = "\t".join(str(x) if x is not None else "" for x in row)
                text_chunks.append(row_str)
        return "\n".join(text_chunks)
    except Exception as e:
        logger.error("Error reading Excel file %s: %s", filepath, e)
        return ""


def _read_ipynb(filepath):
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            data = json.load(f)
        text_cells = []
        for cell in data.get("cells", []):
            # Combine source lines into a single text block
            cell_src = "".join(cell.get("source", []))
            text_cells.append(cell_src)
        return "\n".join(text_cells)
    except Exception as e:
        logger.error("Error reading IPYNB %s: %s", filepath, e)
        return ""
